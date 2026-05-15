from __future__ import annotations

"""Экспорт итогов моделирования в книгу Excel."""

import logging
import math
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

__all__ = ["export_xlsx"]


HEADER_FILL = "BDD7EE"
UNITS_FILL = "D9D9D9"
EMERGENCY_FILL = "FCE4D6"
GREEN_FILL = "E2F0D9"
RED_FILL = "F4CCCC"
MEDIUM_GREY = "808080"
THROAT_COMMENT = "Коэф. > 1.0 означает наличие очереди на горловину (перегрузка)"

HEADER_COMMENTS = {
    "scenario": "Внутренний машинный код сценария, используемый в результатах моделирования.",
    "display_name": "Человекочитаемое название сценария для итогового отчёта.",
    "trains_total": "Количество поездов, обработанных в сценарии.",
    "throughput_trains_per_hour": "Пропускная способность: число поездов за расчётный интервал, пересчитанное в поезда в час.",
    "headway_avg_s": "Средний интервал попутного следования между отправлениями поездов.",
    "dwell_avg_s": "Среднее время стоянки поездов на станции.",
    "throat_utilization": THROAT_COMMENT,
    "mean_travel_time_s": "Среднее полное время нахождения поезда в модели станции.",
    "mean_wait_times_s": "Среднее время ожидания маршрута по событиям симуляции.",
    "recovery_times": "Время восстановления графика после нарушения движения.",
    "cascade_delays": "Суммарная каскадная задержка поездов после нарушения движения.",
    "packet_integrity_ratio": "Доля пакетов виртуальной сцепки, сохранивших целостность.",
    "max_intrapacket_gap_s": "Максимальный интервал между поездами внутри пакета ВС.",
    "delay_arrive_avg_s": "Среднее опоздание прибытия относительно планового времени.",
    "delay_depart_avg_s": "Среднее опоздание отправления относительно планового времени.",
}


SCENARIO_LABELS = {
    "Baseline": "АБ - базовый режим",
    "Demo-AB": "АБ - штатный пропуск",
    "Demo-VC-A": "ВС - пакет А (2 поезда)",
    "Demo-VC-B": "ВС - пакет Б (2 поезда)",
    "AB-Recovery": "АБ - восстановление после задержки",
    "VC-Recovery": "ВС - восстановление после задержки",
    "VC-Packet-Split": "ВС - аварийное разделение пакета",
}

SUMMARY_ORDER = [
    "Baseline",
    "Demo-AB",
    "Demo-VC-A",
    "Demo-VC-B",
    "AB-Recovery",
    "VC-Recovery",
    "VC-Packet-Split",
]

METRIC_COLUMNS = [
    ("scenario", "Код сценария", "(machine key)", False),
    ("display_name", "Сценарий", "(display name)", False),
    ("trains_total", "Число поездов", "поез.", True),
    ("throughput_trains_per_hour", "Пропускная способность", "поез./ч", True),
    ("headway_avg_s", "Интервал попутного следования (ср.)", "с", True),
    ("dwell_avg_s", "Время стоянки (ср.)", "с", True),
    ("throat_utilization", "Коэф. загрузки горловины", "д.е.", True),
    ("mean_travel_time_s", "Среднее время хода", "с", True),
    ("mean_wait_times_s", "Среднее время ожидания маршрута", "с", True),
    ("recovery_times", "Время восстановления", "с", True),
    ("cascade_delays", "Каскадные задержки", "с", True),
    ("packet_integrity_ratio", "Коэф. целостности пакета ВС", "д.е.", True),
    ("max_intrapacket_gap_s", "Макс. внутрипакетный интервал", "с", True),
    ("delay_arrive_avg_s", "Среднее опоздание прибытия", "с", True),
    ("delay_depart_avg_s", "Среднее опоздание отправления", "с", True),
]

METRIC_ALIASES = {
    "mean_wait_times_s": ("mean_wait_time_s",),
    "recovery_times": ("recovery_time_s",),
    "cascade_delays": ("cascade_delay_s",),
    "max_intrapacket_gap_s": ("max_intra_packet_gap_s",),
}

INTEGER_KEYS = {"trains_total"}
THROUGHPUT_KEYS = {"throughput_trains_per_hour"}
LOWER_IS_BETTER_KEYS = {
    "headway_avg_s",
    "dwell_avg_s",
    "throat_utilization",
    "mean_travel_time_s",
    "mean_wait_times_s",
    "recovery_times",
    "cascade_delays",
    "max_intrapacket_gap_s",
    "delay_arrive_avg_s",
    "delay_depart_avg_s",
}
HIGHER_IS_BETTER_KEYS = {"packet_integrity_ratio"}


def export_xlsx(results: dict[str, dict[str, float]], output_dir: Path | str) -> None:
    """Сохраняет сводные результаты моделирования в единый Excel-файл."""
    openpyxl = _load_openpyxl()
    workbook = openpyxl.Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Сводная таблица"
    methodology_sheet = workbook.create_sheet("Методология")
    charts_sheet = workbook.create_sheet("Графики")

    _write_summary_sheet(summary_sheet, results, openpyxl)
    _write_methodology_sheet(methodology_sheet, results, openpyxl)
    _write_charts_sheet(charts_sheet, Path(output_dir), openpyxl)

    path = Path(output_dir) / "simulation_results.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    logger.info("Excel-отчёт сформирован: %s", path.absolute())


def _load_openpyxl() -> Any:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        if exc.name == "openpyxl":
            raise ImportError(
                "Для экспорта Excel требуется зависимость openpyxl>=3.1. "
                "Установите её через requirements.txt."
            ) from exc
        raise

    return _OpenPyXL(
        Workbook=Workbook,
        Comment=Comment,
        Image=Image,
        Alignment=Alignment,
        Border=Border,
        Font=Font,
        PatternFill=PatternFill,
        Side=Side,
        get_column_letter=get_column_letter,
    )


class _OpenPyXL:
    def __init__(self, **items: Any) -> None:
        self.__dict__.update(items)


def _write_summary_sheet(ws: Any, results: dict[str, dict[str, float]], openpyxl: Any) -> None:
    _write_header(ws, openpyxl)
    ws.freeze_panes = "A3"

    row_idx = 3
    for scenario in _ordered_scenarios(results):
        _write_metric_row(ws, row_idx, scenario, results[scenario], openpyxl)
        if scenario == "VC-Packet-Split":
            _fill_row(ws, row_idx, EMERGENCY_FILL, openpyxl)
        _apply_throat_overload_style(ws, row_idx, openpyxl)
        row_idx += 1

    for col_idx, (metric_key, _, _, _) in enumerate(METRIC_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx).comment = openpyxl.Comment(HEADER_COMMENTS[metric_key], "PIMPS")
    _auto_fit_columns(ws, openpyxl)


def _write_methodology_sheet(ws: Any, results: dict[str, dict[str, float]], openpyxl: Any) -> None:
    _write_header(ws, openpyxl)
    ws.freeze_panes = "A3"

    rows: list[tuple[str, str | None, str | None]] = [
        ("source", "Baseline", None),
        ("source", "Demo-VC-A", None),
        ("delta", "Demo-VC-A", "Baseline"),
        ("blank", None, None),
        ("source", "AB-Recovery", None),
        ("source", "VC-Recovery", None),
        ("delta", "VC-Recovery", "AB-Recovery"),
        ("blank", None, None),
        ("source", "VC-Packet-Split", None),
        ("delta", "VC-Packet-Split", "Demo-VC-A"),
    ]

    block_starts = {3, 7, 11}
    row_idx = 3
    for kind, scenario, base_scenario in rows:
        if kind == "blank":
            row_idx += 1
            continue

        if kind == "source" and scenario is not None:
            metrics = results.get(scenario, {})
            _write_metric_row(ws, row_idx, scenario, metrics, openpyxl)
            _apply_throat_overload_style(ws, row_idx, openpyxl)
            if scenario == "VC-Packet-Split":
                ws.cell(row=row_idx, column=1).comment = openpyxl.Comment(
                    "Аварийный режим - разделение пакета ВС",
                    "PIMPS",
                )
        elif kind == "delta" and scenario is not None and base_scenario is not None:
            _write_delta_row(ws, row_idx, scenario, base_scenario, results, openpyxl)

        if row_idx in block_starts:
            _apply_block_top_border(ws, row_idx, openpyxl)
        row_idx += 1

    _auto_fit_columns(ws, openpyxl)


def _write_charts_sheet(ws: Any, output_dir: Path, openpyxl: Any) -> None:
    current_row = 1
    image_specs = [
        ("methodology_comparison_chart*.png", False, None),
        ("occupancy_Demo-VC-A*.png", False, None),
        ("occupancy_Demo-VC-A*.jpg", False, None),
        ("occupancy_Demo-VC-B*.png", False, None),
        ("occupancy_Demo-VC-B*.jpg", False, None),
    ]

    for pattern, single, label in image_specs:
        current_row = _insert_images_for_pattern(ws, output_dir, pattern, current_row, openpyxl, single, label)

    profile_label = "Тяговый профиль: АБ-Recovery, Freight-401, маршрут N-2P"
    profile_images = _find_images(output_dir, "profile_AB-Recovery_Freight-401*.png")
    profile_images.extend(_find_images(output_dir, "profile_AB-Recovery_Freight-401*.jpg"))
    profile_images = sorted(set(profile_images))
    if not profile_images:
        logger.warning("Изображение профиля тяги не найдено: profile_AB-Recovery_Freight-401*.png/jpg")
    else:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        label_cell = ws.cell(row=current_row, column=1, value=profile_label)
        label_cell.font = openpyxl.Font(bold=True)
        current_row += 1
        _insert_image(ws, profile_images[0], f"A{current_row}", openpyxl)


def _insert_images_for_pattern(
    ws: Any,
    output_dir: Path,
    pattern: str,
    current_row: int,
    openpyxl: Any,
    single: bool,
    label: str | None,
) -> int:
    images = _find_images(output_dir, pattern)
    if not images:
        logger.warning("Изображения не найдены по шаблону: %s", pattern)
        return current_row

    selected_images = images[:1] if single else images
    for image_path in selected_images:
        if label:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = openpyxl.Font(bold=True)
            current_row += 1
        height = _insert_image(ws, image_path, f"A{current_row}", openpyxl)
        current_row += max(2, math.ceil(height / 20) + 2)
    return current_row


def _insert_image(ws: Any, image_path: Path, anchor: str, openpyxl: Any) -> int:
    image = openpyxl.Image(str(image_path))
    original_width = image.width
    original_height = image.height
    if original_width > 0:
        ratio = 600 / original_width
        image.width = 600
        image.height = int(original_height * ratio)
    ws.add_image(image, anchor)
    return int(image.height)


def _find_images(output_dir: Path, pattern: str) -> list[Path]:
    if not output_dir.exists():
        logger.warning("Каталог вывода не найден: %s", output_dir)
        return []
    return sorted(path for path in output_dir.rglob(pattern) if path.is_file())


def _write_header(ws: Any, openpyxl: Any) -> None:
    header_fill = openpyxl.PatternFill("solid", fgColor=HEADER_FILL)
    units_fill = openpyxl.PatternFill("solid", fgColor=UNITS_FILL)
    header_font = openpyxl.Font(bold=True, size=11)
    units_font = openpyxl.Font(italic=True, size=10)

    for col_idx, (_, header, unit, _) in enumerate(METRIC_COLUMNS, start=1):
        header_cell = ws.cell(row=1, column=col_idx, value=header)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = openpyxl.Alignment(horizontal="center", vertical="center", wrap_text=True)

        units_cell = ws.cell(row=2, column=col_idx, value=unit)
        units_cell.fill = units_fill
        units_cell.font = units_font
        units_cell.alignment = openpyxl.Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_metric_row(
    ws: Any,
    row_idx: int,
    scenario: str,
    metrics: dict[str, float],
    openpyxl: Any,
) -> None:
    for col_idx, (metric_key, _, _, is_numeric) in enumerate(METRIC_COLUMNS, start=1):
        if metric_key == "scenario":
            value: Any = scenario
        elif metric_key == "display_name":
            value = SCENARIO_LABELS.get(scenario, scenario)
        else:
            value = _metric_value(metrics, metric_key)
        _write_cell(ws, row_idx, col_idx, value, metric_key, is_numeric)


def _write_delta_row(
    ws: Any,
    row_idx: int,
    scenario: str,
    base_scenario: str,
    results: dict[str, dict[str, float]],
    openpyxl: Any,
) -> None:
    label = "Δ к штатному ВС" if scenario == "VC-Packet-Split" else "Δ ВС−АБ"
    ws.cell(row=row_idx, column=1, value=label)
    ws.cell(row=row_idx, column=2, value=f"{SCENARIO_LABELS.get(scenario, scenario)} − {SCENARIO_LABELS.get(base_scenario, base_scenario)}")

    scenario_metrics = results.get(scenario, {})
    base_metrics = results.get(base_scenario, {})
    for col_idx, (metric_key, _, _, is_numeric) in enumerate(METRIC_COLUMNS[2:], start=3):
        if not is_numeric:
            continue
        value = _delta_value(scenario_metrics, base_metrics, metric_key)
        cell = _write_cell(ws, row_idx, col_idx, value, metric_key, True)
        _apply_delta_style(cell, metric_key, value, openpyxl)


def _write_cell(
    ws: Any,
    row_idx: int,
    col_idx: int,
    value: Any,
    metric_key: str,
    is_numeric: bool,
) -> Any:
    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    if is_numeric and isinstance(value, int | float):
        if metric_key in INTEGER_KEYS:
            cell.number_format = "0"
        else:
            cell.number_format = "0.00"
    return cell


def _metric_value(metrics: dict[str, float], metric_key: str) -> float | int | None:
    keys = (metric_key, *METRIC_ALIASES.get(metric_key, ()))
    for key in keys:
        if key not in metrics:
            continue
        value = metrics[key]
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int | float):
            value = float(value)
            if math.isfinite(value):
                return int(value) if metric_key in INTEGER_KEYS else value
            return None
    return None


def _delta_value(
    metrics: dict[str, float],
    base_metrics: dict[str, float],
    metric_key: str,
) -> float | int | None:
    current = _metric_value(metrics, metric_key)
    base = _metric_value(base_metrics, metric_key)
    if current is None or base is None:
        return None
    return float(current) - float(base)


def _ordered_scenarios(results: dict[str, dict[str, float]]) -> list[str]:
    ordered = [scenario for scenario in SUMMARY_ORDER if scenario in results]
    ordered.extend(scenario for scenario in results if scenario not in ordered)
    return ordered


def _column_index(metric_key: str) -> int:
    for idx, (key, _, _, _) in enumerate(METRIC_COLUMNS, start=1):
        if key == metric_key:
            return idx
    raise KeyError(metric_key)


def _fill_row(ws: Any, row_idx: int, color: str, openpyxl: Any) -> None:
    fill = openpyxl.PatternFill("solid", fgColor=color)
    for col_idx in range(1, len(METRIC_COLUMNS) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _apply_throat_overload_style(ws: Any, row_idx: int, openpyxl: Any) -> None:
    throat_col = _column_index("throat_utilization")
    cell = ws.cell(row=row_idx, column=throat_col)
    if isinstance(cell.value, int | float) and cell.value >= 1.0:
        cell.fill = openpyxl.PatternFill("solid", fgColor=EMERGENCY_FILL)
        cell.font = openpyxl.Font(bold=True)


def _apply_delta_style(cell: Any, metric_key: str, value: float | int | None, openpyxl: Any) -> None:
    if value is None or value == 0:
        return

    is_improvement = False
    is_degradation = False
    if metric_key in THROUGHPUT_KEYS or metric_key in HIGHER_IS_BETTER_KEYS:
        is_improvement = value > 0
        is_degradation = value < 0
    elif metric_key in LOWER_IS_BETTER_KEYS:
        is_improvement = value < 0
        is_degradation = value > 0

    if is_improvement:
        cell.fill = openpyxl.PatternFill("solid", fgColor=GREEN_FILL)
    elif is_degradation:
        cell.fill = openpyxl.PatternFill("solid", fgColor=RED_FILL)


def _apply_block_top_border(ws: Any, row_idx: int, openpyxl: Any) -> None:
    side = openpyxl.Side(style="medium", color=MEDIUM_GREY)
    for col_idx in range(1, len(METRIC_COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = openpyxl.Border(
            left=cell.border.left,
            right=cell.border.right,
            top=side,
            bottom=cell.border.bottom,
        )


def _auto_fit_columns(ws: Any, openpyxl: Any) -> None:
    for col_idx in range(1, len(METRIC_COLUMNS) + 1):
        letter = openpyxl.get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(40, max(15, max_len + 2))
