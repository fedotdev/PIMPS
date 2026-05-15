from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image as PilImage

from src.renderers.excel_reporter import export_xlsx


def _metrics(
    trains_total: int,
    throughput: float,
    throat_utilization: float,
    recovery_time: float,
) -> dict[str, float]:
    return {
        "trains_total": float(trains_total),
        "throughput_trains_per_hour": throughput,
        "headway_avg_s": 120.0,
        "dwell_avg_s": 60.0,
        "throat_utilization": throat_utilization,
        "mean_travel_time_s": 300.0,
        "mean_wait_time_s": 20.0,
        "recovery_time_s": recovery_time,
        "cascade_delay_s": 15.0,
        "packet_integrity_ratio": 1.0,
        "max_intra_packet_gap_s": 45.0,
        "delay_arrive_avg_s": 5.0,
        "delay_depart_avg_s": 8.0,
    }


def _sample_results() -> dict[str, dict[str, float]]:
    return {
        "Baseline": _metrics(2, 12.0, 1.10, 0.0),
        "Demo-AB": _metrics(8, 10.0, 0.80, 0.0),
        "Demo-VC-A": _metrics(2, 16.0, 0.90, 0.0),
        "Demo-VC-B": _metrics(2, 18.0, 0.95, 0.0),
        "AB-Recovery": _metrics(6, 8.0, 1.20, 320.0),
        "VC-Recovery": _metrics(6, 14.0, 1.30, 180.0),
        "VC-Packet-Split": _metrics(3, 9.0, 1.40, 0.0),
    }


def _column_by_header(sheet, header: str) -> int:
    for cell in sheet[1]:
        if cell.value == header:
            return int(cell.column)
    raise AssertionError(f"Колонка не найдена: {header}")


def _row_by_code(sheet, code: str) -> int:
    for row_idx in range(3, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == code:
            return row_idx
    raise AssertionError(f"Строка не найдена: {code}")


def _rgb(cell) -> str:
    return str(cell.fill.fgColor.rgb)


def test_export_xlsx_summary_and_methodology_layout(tmp_path: Path) -> None:
    export_xlsx(_sample_results(), tmp_path)

    workbook = load_workbook(tmp_path / "simulation_results.xlsx")
    assert workbook.sheetnames == ["Сводная таблица", "Методология", "Графики"]

    summary = workbook["Сводная таблица"]
    assert summary.freeze_panes == "A3"
    assert summary.cell(row=1, column=1).value == "Код сценария"
    assert summary.cell(row=2, column=1).value == "(machine key)"

    throat_col = _column_by_header(summary, "Коэф. загрузки горловины")
    assert summary.cell(row=1, column=throat_col).comment.text == (
        "Коэф. > 1.0 означает наличие очереди на горловину (перегрузка)"
    )
    baseline_row = _row_by_code(summary, "Baseline")
    assert _rgb(summary.cell(row=baseline_row, column=throat_col)).endswith("FCE4D6")
    assert summary.cell(row=baseline_row, column=throat_col).font.bold

    split_row = _row_by_code(summary, "VC-Packet-Split")
    assert _rgb(summary.cell(row=split_row, column=1)).endswith("FCE4D6")

    methodology = workbook["Методология"]
    assert methodology.freeze_panes == "A3"
    assert methodology.cell(row=2, column=1).value == "(machine key)"

    source_codes = {
        methodology.cell(row=row_idx, column=1).value
        for row_idx in range(3, methodology.max_row + 1)
        if methodology.cell(row=row_idx, column=1).value in _sample_results()
    }
    assert source_codes == set(_sample_results())

    delta_rows = [
        row_idx
        for row_idx in range(3, methodology.max_row + 1)
        if str(methodology.cell(row=row_idx, column=1).value).startswith("Δ")
    ]
    assert len(delta_rows) == 3

    blank_rows = [
        row_idx
        for row_idx in range(3, methodology.max_row + 1)
        if all(methodology.cell(row=row_idx, column=col_idx).value is None for col_idx in range(1, 16))
    ]
    assert len(blank_rows) == 2

    recovery_col = _column_by_header(methodology, "Время восстановления")
    vc_recovery_row = _row_by_code(methodology, "VC-Recovery")
    recovery_delta_row = vc_recovery_row + 1
    assert methodology.cell(row=recovery_delta_row, column=1).value == "Δ ВС−АБ"
    assert methodology.cell(row=recovery_delta_row, column=recovery_col).value == -140.0
    assert _rgb(methodology.cell(row=recovery_delta_row, column=recovery_col)).endswith("E2F0D9")


def test_export_xlsx_embeds_recovery_profile_image(tmp_path: Path) -> None:
    profile_dir = tmp_path / "profiles"
    profile_dir.mkdir()
    profile_path = profile_dir / "profile_AB-Recovery_Freight-401_route_N_2P.png"
    PilImage.new("RGB", (10, 10), "white").save(profile_path)

    export_xlsx(_sample_results(), tmp_path)

    workbook = load_workbook(tmp_path / "simulation_results.xlsx")
    charts = workbook["Графики"]
    assert any(
        cell.value == "Тяговый профиль: АБ-Recovery, Freight-401, маршрут N-2P"
        for row in charts.iter_rows()
        for cell in row
    )
    assert len(charts._images) == 1
